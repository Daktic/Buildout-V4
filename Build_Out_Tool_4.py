import tkinter as tk
from tkinter import IntVar
from openpyxl import Workbook

root= tk.Tk()

canvas1 = tk.Canvas(root, width = 1700, height = 800)
canvas1.pack()
root.attributes("-fullscreen", True)





keywords = tk.Entry (root) 
branded_words = tk.Entry (root) 
#brand
tactic_type = tk.Entry (root)
tier = tk.Entry (root)
corp_non = tk.Entry (root)
lob = tk.Entry (root)
intiative = tk.Entry (root)
#match_type
region = tk.Entry (root)
market = tk.Entry (root)
sub_region = tk.Entry (root)
language = tk.Entry (root)
device = tk.Entry (root)

topic = tk.Entry(root)

url_google = tk.Entry (root)
url_bing = tk.Entry (root)

url_g_aw = tk.Entry (root)
url_g_it = tk.Entry (root)
url_g_co = tk.Entry (root)
url_g_ev = tk.Entry (root)

url_b_aw = tk.Entry (root)
url_b_it = tk.Entry (root)
url_b_co = tk.Entry (root)
url_b_ev = tk.Entry (root)

#cj_stages

canvas1.create_window(100, 140, window=keywords)
canvas1.create_window(350, 140, window=branded_words)
canvas1.create_window(250, 240, window=tactic_type)
canvas1.create_window(400, 240, window=tier)
canvas1.create_window(550, 240, window=corp_non)
canvas1.create_window(700, 240, window=lob)
canvas1.create_window(850, 240, window=intiative)
canvas1.create_window(1000, 240, window=region)
canvas1.create_window(1150, 240, window=market)
canvas1.create_window(1300, 240, window=sub_region)
canvas1.create_window(1450, 240, window=language)
canvas1.create_window(1600, 240, window=device)
canvas1.create_window(600, 140, window=topic)
canvas1.create_window(100, 340, window=url_google)
canvas1.create_window(250, 340, window=url_bing)

canvas1.create_window(700, 440, window=url_g_aw)
canvas1.create_window(850, 440, window=url_b_aw)
canvas1.create_window(700, 540, window=url_g_it)
canvas1.create_window(850, 540, window=url_b_it)
canvas1.create_window(700, 640, window=url_g_co)
canvas1.create_window(850, 640, window=url_b_co)
canvas1.create_window(700, 740, window=url_g_ev)
canvas1.create_window(850, 740, window=url_b_ev)


keyword_label = tk.Label(root, text='Keywords with "/" delimintor')
branded_words_label = tk.Label(root, text='Branded Words with "/" delimintor')
tactic_type_label = tk.Label(root, text='AW/DG')
tier_label = tk.Label(root, text='Tier')
corp_non_label = tk.Label(root, text='Corp/Non-Corp')
lob_label = tk.Label(root, text='L.O.B')
intiative_label = tk.Label(root, text='Intitative')
region_label = tk.Label(root, text='Region')
market_label = tk.Label(root, text='Market')
sub_region_label = tk.Label(root, text='Sub Region')
language_label = tk.Label(root, text='Language')
device_label = tk.Label(root, text='Device')
topic_label = tk.Label(root, text='Topic')
url_google_label = tk.Label(root, text='URL Google')
url_bing_label = tk.Label(root, text='URL Bing')

cj_url_labels = tk.Label(root, text='URLS for CJ stage')
google_label = tk.Label(root, text='Google')
bing_label = tk.Label(root, text='Bing')
aw_label = tk.Label(root, text='AW')
it_label = tk.Label(root, text='IT')
co_label = tk.Label(root, text='CO')
ev_label = tk.Label(root, text='EV')




canvas1.create_window(100, 100, window=keyword_label)
canvas1.create_window(350, 100, window=branded_words_label)
canvas1.create_window(600, 100, window=topic_label)
canvas1.create_window(250, 200, window=tactic_type_label)
canvas1.create_window(400, 200, window=tier_label)
canvas1.create_window(550, 200, window=corp_non_label)
canvas1.create_window(700, 200, window=lob_label)
canvas1.create_window(850, 200, window=intiative_label)
canvas1.create_window(1000, 200, window=region_label)
canvas1.create_window(1150, 200, window=market_label)
canvas1.create_window(1300, 200, window=sub_region_label)
canvas1.create_window(1450, 200, window=language_label)
canvas1.create_window(1600, 200, window=device_label)
canvas1.create_window(100, 300, window=url_google_label)
canvas1.create_window(220, 300, window=url_bing_label)

canvas1.create_window(770, 380, window=cj_url_labels)
canvas1.create_window(700, 410, window=google_label)
canvas1.create_window(840, 410, window=bing_label)

canvas1.create_window(600, 440, window=aw_label)
canvas1.create_window(600, 540, window=it_label)
canvas1.create_window(600, 640, window=co_label)
canvas1.create_window(600, 740, window=ev_label)

v = IntVar()
v.set(1)

tk.Radiobutton(root, text="Consumer Journey", variable=v, value=1).pack(anchor='w')
tk.Radiobutton(root, text="Consumer Journey - No Stage", variable=v, value=2).pack(anchor='w')
tk.Radiobutton(root, text="Audiance Pilot", variable=v, value=3).pack(anchor='w')


def getInfo ():  
    build_keyword = keywords.get()
    build_branded_words = branded_words.get().lower()
    build_keyword_split = build_keyword.split('/')
    build_branded_words_split = build_branded_words.split('/')

    build_tactic_type = tactic_type.get().upper()
    build_tier= 'T' + tier.get()
    build_corp_non = corp_non.get()
    build_lob = lob.get()
    build_initative = intiative.get()
    build_region = region.get().upper()
    build_market = market.get().upper()
    build_sub_region = sub_region.get()
    build_language = language.get().upper()
    build_device = device.get()
    build_topic = topic.get()
    build_url_google = url_google.get()
    build_url_bing = url_bing.get()
    
    return build_keyword_split, build_branded_words_split, build_tactic_type, build_tier, build_corp_non, build_lob, build_initative, build_region, build_market, build_sub_region, build_language, build_device, build_topic, build_url_google, build_url_bing
    
def get_cj_URLS():
    aw_url_g = url_g_aw.get()
    aw_url_b = url_b_aw.get()
    it_url_g = url_g_it.get()
    it_url_b = url_b_it.get()
    co_url_g = url_g_co.get()
    co_url_b = url_b_co.get()
    ev_url_g = url_g_ev.get()
    ev_url_b = url_b_ev.get()
    
    return aw_url_g, aw_url_b, it_url_g, it_url_b, co_url_g, co_url_b, ev_url_g, ev_url_b


def taxonomyBuild():
    info = getInfo()
    radio = v.get()
    br = 'BR_'
    ub = 'UB_'
    broad = '_Broad_'
    exact = '_Exact_'
    campaign_br_broad = br +'_'.join(info[2:7]) + broad + '_'.join(info[7:-3])
    campaign_br_exact = br +'_'.join(info[2:7]) + exact + '_'.join(info[7:-3])
    campaign_ub_broad = ub +'_'.join(info[2:7]) + broad + '_'.join(info[7:-3])
    campaign_ub_exact = ub +'_'.join(info[2:7]) + exact + '_'.join(info[7:-3])
    if radio == 3:
        hiq = '_HIQ'
        miq = '_MIQ'
        unk = '_UNK'
        
        adgroup_br_broad_HIQ = br + str(info[5]) + '_' + str(info[2]) + hiq + broad + str(info[7]) + '_' + str(info[10])
        adgroup_br_broad_MIQ = br + str(info[5]) + '_' + str(info[2]) + miq + broad + str(info[7]) + '_' + str(info[10])
        adgroup_br_broad_UNK = br + str(info[5]) + '_' + str(info[2]) + unk + broad + str(info[7]) + '_' + str(info[10])

        adgroup_br_exact_HIQ = br + str(info[5]) + '_' + str(info[2]) + hiq + exact + str(info[7]) + '_' + str(info[10])
        adgroup_br_exact_MIQ = br + str(info[5]) + '_' + str(info[2]) + miq + exact + str(info[7]) + '_' + str(info[10])
        adgroup_br_exact_UNK = br + str(info[5]) + '_' + str(info[2]) + unk + exact + str(info[7]) + '_' + str(info[10])

        adgroup_ub_broad_HIQ = ub + str(info[5]) + '_' + str(info[2]) + hiq + broad + str(info[7]) + '_' + str(info[10])
        adgroup_ub_broad_MIQ = ub + str(info[5]) + '_' + str(info[2]) + miq + broad + str(info[7]) + '_' + str(info[10])
        adgroup_ub_broad_UNK = ub + str(info[5]) + '_' + str(info[2]) + unk + broad + str(info[7]) + '_' + str(info[10])

        adgroup_ub_exact_HIQ = ub + str(info[5]) + '_' + str(info[2]) + hiq + exact + str(info[7]) + '_' + str(info[10])
        adgroup_ub_exact_MIQ = ub + str(info[5]) + '_' + str(info[2]) + miq + exact + str(info[7]) + '_' + str(info[10])
        adgroup_ub_exact_UNK = ub + str(info[5]) + '_' + str(info[2]) + unk + exact + str(info[7]) + '_' + str(info[10])

        return campaign_br_broad, campaign_br_exact, campaign_ub_broad, campaign_ub_exact, adgroup_br_broad_HIQ, adgroup_br_broad_MIQ, adgroup_br_broad_UNK, adgroup_br_exact_HIQ, adgroup_br_exact_MIQ, adgroup_br_exact_UNK, adgroup_ub_broad_HIQ,adgroup_ub_broad_MIQ, adgroup_ub_broad_UNK, adgroup_ub_exact_HIQ,adgroup_ub_exact_MIQ, adgroup_ub_exact_UNK
    elif radio == 1:
        aw = '_AW'
        it = '_IT'
        co = '_CO'
        ev = '_EV'
        adgroup_br_broad_AW = br + str(info[6]) + '_' + str(info[-3]) +aw +broad + str(info[-5])
        adgroup_br_broad_IT = br + str(info[6]) + '_' + str(info[-3]) +it +broad + str(info[-5])
        adgroup_br_broad_CO = br + str(info[6]) + '_' + str(info[-3]) +co +broad + str(info[-5])
        adgroup_br_broad_EV = br + str(info[6]) + '_' + str(info[-3]) +ev +broad + str(info[-5])

        adgroup_br_exact_AW = br + str(info[6]) + '_' + str(info[-3]) +aw +exact + str(info[-5])
        adgroup_br_exact_IT = br + str(info[6]) + '_' + str(info[-3]) +it +exact + str(info[-5])
        adgroup_br_exact_CO = br + str(info[6]) + '_' + str(info[-3]) +co +exact + str(info[-5])
        adgroup_br_exact_EV = br + str(info[6]) + '_' + str(info[-3]) +ev +exact + str(info[-5])

        adgroup_ub_broad_AW = ub + str(info[6]) + '_' + str(info[-3]) +aw +broad + str(info[-5])
        adgroup_ub_broad_IT = ub + str(info[6]) + '_' + str(info[-3]) +it +broad + str(info[-5])
        adgroup_ub_broad_CO = ub + str(info[6]) + '_' + str(info[-3]) +co +broad + str(info[-5])
        adgroup_ub_broad_EV = ub + str(info[6]) + '_' + str(info[-3]) +ev +broad + str(info[-5])

        adgroup_ub_exact_AW = ub + str(info[6]) + '_' + str(info[-3]) +aw +exact + str(info[-5])
        adgroup_ub_exact_IT = ub + str(info[6]) + '_' + str(info[-3]) +it +exact + str(info[-5])
        adgroup_ub_exact_CO = ub + str(info[6]) + '_' + str(info[-3]) +co +exact + str(info[-5])
        adgroup_ub_exact_EV = ub + str(info[6]) + '_' + str(info[-3]) +ev +exact + str(info[-5])

        return campaign_br_broad, campaign_br_exact, campaign_ub_broad, campaign_ub_exact, adgroup_br_broad_AW, adgroup_br_broad_IT, adgroup_br_broad_CO, adgroup_br_broad_EV, adgroup_br_exact_AW, adgroup_br_exact_IT, adgroup_br_exact_CO, adgroup_br_exact_EV, adgroup_ub_broad_AW, adgroup_ub_broad_IT, adgroup_ub_broad_CO, adgroup_ub_broad_EV, adgroup_ub_exact_AW, adgroup_ub_exact_IT, adgroup_ub_exact_CO, adgroup_ub_exact_EV


    else:
        adgroup_br_broad = br + str(info[6]) + '_' + str(info[-3]) +broad + str(info[-8]) + '_'  + str(info[-5])
        adgroup_br_exact = br + str(info[6]) + '_' + str(info[-3]) +exact + str(info[-8]) + '_'  + str(info[-5])
        adgroup_ub_broad = ub + str(info[6]) + '_' + str(info[-3]) +broad + str(info[-8]) + '_'  + str(info[-5])
        adgroup_ub_exact = ub + str(info[6]) + '_' + str(info[-3]) +exact + str(info[-8]) + '_'  + str(info[-5])

        return campaign_br_broad, campaign_br_exact, campaign_ub_broad, campaign_ub_exact, adgroup_br_broad, adgroup_br_exact, adgroup_ub_broad, adgroup_ub_exact

def createBuild():
    taxonomy = taxonomyBuild()
    info = getInfo()
    urls = get_cj_URLS()

    wb = Workbook()
    ws = wb.active
    radio = v.get()
    if radio == 1:
        #root words
        pre_rootword_what ='What'
        pre_rootword_how ='How'
        pre_rootword_explain ='Explain'
        pre_rootword_best ='Best'
        pre_rootword_top ='Top'
        pre_rootword_compare ='Compare'
        pre_rootword_try ='Try'
        post_rootword_explained ='Explained'
        post_rootword_examples ='Examples'
        post_rootword_sources ='Sources'
        post_rootword_versus ='Versus'
        post_rootword_review ='Review'
        post_rootword_trial ='Trial'
        post_rootword_demo ='Demo'

        #create list permiations from rootwords and user generated keywords
            #Awareness
        Awareness_perm_what = [pre_rootword_what + ' ' + x for x in info[0]]
        Awareness_perm_how = [pre_rootword_how + ' ' + x for x in info[0]]
        Awareness_perm_explain = [pre_rootword_explain + ' ' + x for x in info[0]]
        Awareness_perm_explained = [x + ' ' + post_rootword_explained for x in info[0]]
            #Interest
        Interest_perm_best = [pre_rootword_best + ' ' + x for x in info[0]]
        Interest_perm_top = [pre_rootword_top + ' ' + x for x in info[0]]
        Interest_perm_examples = [x + ' ' + post_rootword_examples for x in info[0]]
        Interest_perm_sources = [x + ' ' + post_rootword_sources for x in info[0]]
            #Consideration
        Consideration_perm_compare = [pre_rootword_compare + ' ' + x for x in info[0]]
        Consideration_perm_versus  = [x + ' ' + post_rootword_versus for x in info[0]]
        Consideration_perm_review  = [x + ' ' + post_rootword_review for x in info[0]]
            #Evaluation
        Evaluation_perm_try = [pre_rootword_try + ' ' + x for x in info[0]]
        Evaluation_perm_trial = [x + ' ' + post_rootword_trial for x in info[0]]
        Evaluation_perm_demo = [x + ' ' + post_rootword_demo for x in info[0]]

        #Grouping by Stage
        Awareness = Awareness_perm_what + Awareness_perm_how + Awareness_perm_explain + Awareness_perm_explained
        Interest = Interest_perm_best + Interest_perm_top + Interest_perm_examples + Interest_perm_sources
        Consideration = Consideration_perm_compare + Consideration_perm_versus + Consideration_perm_review
        Evaluation = Evaluation_perm_try + Evaluation_perm_trial + Evaluation_perm_demo

        #creates broad match keywords
        Awareness_broad_1 = ["+" + suit for suit in Awareness]
        Awareness_broad_2 = [x.replace(' ',' +') for x in Awareness_broad_1]
        Awareness = Awareness + Awareness_broad_2
        Interest_broad_1= ["+" + suit for suit in Interest]
        Interest_broad_2 = [x.replace(' ',' +') for x in Interest_broad_1]
        Interest = Interest + Interest_broad_2
        Consideration_broad_1 = ["+" + suit for suit in Consideration]
        Consideration_broad_2 = [x.replace(' ', ' +') for x in Consideration_broad_1]
        Consideration = Consideration + Consideration_broad_2
        Evaluation_broad_1 = ["+" + suit for suit in Evaluation]
        Evaluation_broad_2 = [x.replace(' ', ' +') for x in Evaluation_broad_1]
        Evaluation = Evaluation + Evaluation_broad_2
        consumer_journey_stages =[Awareness,Interest, Consideration, Evaluation]
            

        #create worksheets
    
        ws1 = ws.title = "Awareness"
        ws2 = wb.create_sheet("Interest",1)
        ws3 = wb.create_sheet("Consideration",2)
        ws4 = wb.create_sheet("Evaluation",3)
        def column_title(worksheet_name):
            for ws in wb:
                ws['A1'] = 'Campaign'
                ws['B1'] = 'Ad Group'
                ws['C1'] = 'Keyword'
                ws['D1'] = 'Match Type'
                ws['E1'] = 'Max CPC'
                ws['F1'] = 'Final URL Google'
                ws['G1'] = 'Final URL Bing'
                ws['H1'] = 'BR/UB'

        #name each column
        column_title(wb["Awareness"])
        column_title(wb["Interest"])
        column_title(wb["Consideration"])
        column_title(wb["Evaluation"])

        #keyword fill function
        def kfill(worksheet_name, journey_stage):
            ws = worksheet_name   
            r=2
            for word in journey_stage:
                ws.cell(row=r, column=3).value = word
                r+=1

            
        #fill keyword sett to column
        kfill(wb["Awareness"], Awareness)
        kfill(wb["Interest"], Interest)
        kfill(wb["Consideration"], Consideration)
        kfill(wb["Evaluation"], Evaluation)

        #match type/Max CPC/BN/UB fill function
        def mfill(worksheet_name):
            ws = worksheet_name
            r = 2 
            #iterates through rows and adds formula
                #Match Type
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if '+' in ws.cell(row=r, column=3).value:
                        ws.cell(row=r, column=4).value = 'Broad'
                    else:
                        ws.cell(row=r, column=4).value = 'Exact'
                    if ws.cell(row=r, column=4).value =='Broad':
                        ws.cell(row=r, column=5).value = 10
                    else:
                        ws.cell(row=r, column=5).value = 12
                    if any(word.lower() in ws.cell(row=r, column=3).value for word in info[1]):
                        ws.cell(row=r, column=8).value = 'BR'
                    else:
                        ws.cell(row=r, column=8).value = 'UB'
                    r += 1
                
                        
        mfill(wb["Awareness"])
        mfill(wb["Interest"])
        mfill(wb["Consideration"])
        mfill(wb["Evaluation"])

        #URL fill function
        def ufill(worksheet_name,url_asset_G,url_asset_B):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    ws.cell(row=r, column=6).value = str(url_asset_G)
                    ws.cell(row=r, column=7).value = str(url_asset_B)
                r += 1
                
        ufill(wb["Awareness"],urls[0], urls[1])
        ufill(wb["Interest"],urls[2], urls[3])
        ufill(wb["Consideration"],urls[4], urls[5])
        ufill(wb["Evaluation"],urls[6],urls[7])

        def afill(worksheet_name):
            r = 2
            ws = worksheet_name
            
            if ws == wb["Awareness"]:
                for cell in ws:
                    if ws.cell(row=r, column=3).value is not None:
                        if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[4]
                        elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                            ws.cell(row=r, column=2).value = taxonomy[12]
                        elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[8]
                        else:
                            ws.cell(row=r, column=2).value = taxonomy[16]
                        r+=1
            elif ws == wb["Interest"]:
                for cell in ws:
                    if ws.cell(row=r, column=3).value is not None:
                        if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[5]
                        elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                            ws.cell(row=r, column=2).value = taxonomy[13]
                        elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[9]
                        else:
                            ws.cell(row=r, column=2).value = taxonomy[17]
                        r+=1
            elif ws == wb["Consideration"]:
                for cell in ws:
                    if ws.cell(row=r, column=3).value is not None:
                        if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BN':
                            ws.cell(row=r, column=2).value = taxonomy[6]
                        elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                            ws.cell(row=r, column=2).value = taxonomy[14]
                        elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[10]
                        else:
                            ws.cell(row=r, column=2).value = taxonomy[18]
                        r+=1
            elif ws == wb["Evaluation"]:
                for cell in ws:
                    if ws.cell(row=r, column=3).value is not None:
                        if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[7]
                        elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                            ws.cell(row=r, column=2).value = taxonomy[15]
                        elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                            ws.cell(row=r, column=2).value = taxonomy[11]
                        else:
                            ws.cell(row=r, column=2).value = taxonomy[19]
                        r+=1


        afill(wb["Awareness"])
        afill(wb["Interest"])
        afill(wb["Consideration"])
        afill(wb["Evaluation"])

        def cfill(worksheet_name):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                        ws.cell(row=r, column=1).value = taxonomy[0]
                    elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                        ws.cell(row=r, column=1).value = taxonomy[2]
                    elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                        ws.cell(row=r, column=1).value = taxonomy[1]
                    else:
                        ws.cell(row=r, column=1).value = taxonomy[3]
                r +=1

        cfill(wb["Awareness"])
        cfill(wb["Interest"])
        cfill(wb["Consideration"])
        cfill(wb["Evaluation"])

        #Export to Excell
        wb.save(filename="Build Out_CJ_Stage.xlsx")
    
    if radio == 2:
        ws1 = ws.title = "Build"
        

        def column_title(worksheet_name):
            for ws in wb:
                ws['A1'] = 'Campaign'
                ws['B1'] = 'Ad Group'
                ws['C1'] = 'Keyword'
                ws['D1'] = 'Match Type'
                ws['E1'] = 'Max CPC'
                ws['F1'] = 'Final URL Google'
                ws['G1'] = 'Final URL Bing'
                ws['H1'] = 'BN/UB'

        #name each column
        column_title(wb["Build"])

        #keyword fill function
        def kfill(worksheet_name):
            ws = worksheet_name 

            broad_1 = ["+" + suit for suit in info[0]]
            broad_2 = [x.replace(' ',' +') for x in broad_1]
            
            keywords_all = info[0] + broad_2

            r=2
            for word in keywords_all:
                ws.cell(row=r, column=3).value = word
                r+=1
        #fill keyword sett to column
        kfill(wb["Build"])

        #match type/Max CPC/BN/UB fill function
        def mfill(worksheet_name):
            ws = worksheet_name
            r = 2 
            #iterates through rows and adds formula
                #Match Type
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if '+' in ws.cell(row=r, column=3).value:
                        ws.cell(row=r, column=4).value = 'Broad'
                    else:
                        ws.cell(row=r, column=4).value = 'Exact'
                    if ws.cell(row=r, column=4).value =='Broad':
                        ws.cell(row=r, column=5).value = 5
                    else:
                        ws.cell(row=r, column=5).value = 8
                    if any(word.lower() in ws.cell(row=r, column=3).value for word in info[1]):
                        ws.cell(row=r, column=8).value = 'BN'
                    else:
                        ws.cell(row=r, column=8).value = 'UB'
                    r += 1
                
                        
        mfill(wb["Build"])

        #URL fill function
        def ufill(worksheet_name,url_g, url_b):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    ws.cell(row=r, column=6).value = str(info[-2])
                    ws.cell(row=r, column=7).value = str(info[-1])
                r += 1
                
        ufill(wb["Build"],urls[-2], urls[-1])

        def cfill(worksheet_name):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BN':
                        ws.cell(row=r, column=1).value = taxonomy[0]
                    elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                        ws.cell(row=r, column=1).value = taxonomy[2]
                    elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BN':
                        ws.cell(row=r, column=1).value = taxonomy[1]
                    else:
                        ws.cell(row=r, column=1).value = taxonomy[3]
                r +=1

        cfill(wb["Build"])

        def afill(worksheet_name):
            r = 2
            ws = worksheet_name
            
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BN':
                        ws.cell(row=r, column=2).value = taxonomy[4]
                    elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                        ws.cell(row=r, column=2).value = taxonomy[6]
                    elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BN':
                        ws.cell(row=r, column=2).value = taxonomy[5]
                    else:
                        ws.cell(row=r, column=2).value = taxonomy[7]
                r+=1
        afill(wb['Build'])

        wb.save(filename="Build Out_CJ_No_Stage.xlsx")

    if radio ==3:
        ws1 = ws.title = "Build Out"
        
        def column_title(worksheet_name):
            for ws in wb:
                    ws['A1'] = 'Campaign'
                    ws['B1'] = 'Ad Group'
                    ws['C1'] = 'Keyword'
                    ws['D1'] = 'Match Type'
                    ws['E1'] = 'Max CPC'
                    ws['F1'] = 'Final URL Google'
                    ws['G1'] = 'Final URL Bing'
                    ws['H1'] = 'BR/UB'

        column_title(wb["Build Out"])

        #keyword fill function
        def kfill(worksheet_name):
            ws = worksheet_name

            broad_1 = ["+" + suit for suit in info[0]]
            broad_2 = [x.replace(' ',' +') for x in broad_1]
            
            keywords_all = info[0] + broad_2
            
            r=2
            for word in keywords_all:
                for i in range(3):
                    ws.cell(row=r, column=3).value = word
                    r+=1
        kfill(wb["Build Out"])    

        #match type/Max CPC/BN/UB fill function
        def mfill(worksheet_name):
            ws = worksheet_name
            r = 2 
            #iterates through rows and adds formula
                #Match Type
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if '+' in ws.cell(row=r, column=3).value:
                        ws.cell(row=r, column=4).value = 'Broad'
                    else:
                        ws.cell(row=r, column=4).value = 'Exact'
                    if ws.cell(row=r, column=4).value =='Broad':
                        ws.cell(row=r, column=5).value = 8
                    else:
                        ws.cell(row=r, column=5).value = 5
                    oracle = ws.cell(row=r, column=3).value
                    if any((word).lower() in ws.cell(row=r, column=3).value for word in info[1]):
                        ws.cell(row=r, column=8).value = 'BR'
                    else:
                        ws.cell(row=r, column=8).value = 'UB'
                    r += 1
        mfill(wb["Build Out"])  


        def ufill(worksheet_name, landingpage_google, landingpage_bing):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    for i in range(3):
                        ws.cell(row=r, column=6).value = str(info[-2])
                        ws.cell(row=r, column=7).value = str(info[-1])
                        r += 1
        ufill(wb["Build Out"], url_google, url_bing)

        #fill in ad groups
        def afill(worksheet_name):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                        ws.cell(row=r, column=2).value = taxonomy[4]
                        ws.cell(row=r+1, column=2).value = taxonomy[5]
                        ws.cell(row=r+2, column=2).value = taxonomy[6]
                    elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                        ws.cell(row=r, column=2).value = taxonomy[10]
                        ws.cell(row=r+1, column=2).value = taxonomy[11]
                        ws.cell(row=r+2, column=2).value = taxonomy[12]
                    elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                        ws.cell(row=r, column=2).value = taxonomy[7]
                        ws.cell(row=r+1, column=2).value = taxonomy[8]
                        ws.cell(row=r+2, column=2).value = taxonomy[9]
                    else:
                        ws.cell(row=r, column=2).value = taxonomy[13]
                        ws.cell(row=r+1, column=2).value = taxonomy[14]
                        ws.cell(row=r+2, column=2).value = taxonomy[15]
                r+=3
        afill(wb["Build Out"])  

        def cfill(worksheet_name):
            ws = worksheet_name
            r = 2
            for cell in ws:
                if ws.cell(row=r, column=3).value is not None:
                    if ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'BR':
                        for i in range(3):
                            ws.cell(row=r, column=1).value = taxonomy[0]
                    elif ws.cell(row=r, column=4).value == 'Broad' and ws.cell(row=r, column=8).value == 'UB':
                        for i in range(3):
                            ws.cell(row=r, column=1).value = taxonomy[2]
                    elif ws.cell(row=r, column=4).value == 'Exact' and ws.cell(row=r, column=8).value == 'BR':
                        for i in range(3):    
                            ws.cell(row=r, column=1).value = taxonomy[1]
                    else:
                        for i in range(3):
                            ws.cell(row=r, column=1).value = taxonomy[3]
                r +=1
        cfill(wb["Build Out"])


        wb.save(filename="Build Out_Audience_Pilot.xlsx")

def close_window(): 
    root.destroy()

button1 = tk.Button(text='Create Build',fg='green',bg='dark gray', command=createBuild, anchor ='w')
button2 = tk.Button(text='Close window', fg='red', command=close_window, anchor ='w')
button1.pack()
button2.pack()

root.mainloop()